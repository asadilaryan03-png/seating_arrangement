from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
import re
import io
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from datetime import datetime

class ExportManager:
    """Generate PDF and Excel exports for seating arrangements"""
    
    INSTITUTE_NAME = "V. V. P. Institute of Engineering & Technology, Solapur"
    CENTER_CODE = "6321"
    
    @staticmethod
    def generate_pdf(data):
        """Generate PDF with seating arrangements"""
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=0.5*inch, 
                               leftMargin=0.5*inch, topMargin=0.5*inch, bottomMargin=0.5*inch)
        elements = []
        styles = getSampleStyleSheet()
        
        # Custom styles - balanced sizes for readability and fit
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            textColor=colors.black,
            spaceAfter=6,
            alignment=TA_CENTER,
            fontName='Helvetica-Bold'
        )

        subtitle_style = ParagraphStyle(
            'CustomSubtitle',
            parent=styles['Normal'],
            fontSize=16,
            textColor=colors.black,
            spaceAfter=12,
            alignment=TA_CENTER,
            fontName='Helvetica-Bold'
        )

        normal_style = ParagraphStyle(
            'CustomNormal',
            parent=styles['Normal'],
            fontSize=12,
            textColor=colors.black,
            spaceAfter=4
        )
        
        # Add title
        elements.append(Paragraph(ExportManager.INSTITUTE_NAME, title_style))
        elements.append(Spacer(1, 0.1*inch))
        
        # Determine exam type and current date (weekday + date)
        now = datetime.now()
        month = now.month
        year = now.year
        # Institution mapping: Nov-Jan -> Winter, Apr-Jul -> Summer
        if month in [11, 12, 1]:
            exam_label = "Winter"
        elif month in [4, 5, 6, 7]:
            exam_label = "Summer"
        else:
            exam_label = ""
        exam_full = f"{exam_label + ' ' if exam_label else ''}Examination {year}"
        date_str = now.strftime('%A, %d %B %Y')

        # Use styled paragraph (avoid raw HTML tags inside table cells)
        elements.append(Paragraph(exam_full, subtitle_style))
        elements.append(Spacer(1, 0.05*inch))
        # Add seating arrangement title in bold
        elements.append(Paragraph("Seating Arrangement", title_style))
        elements.append(Spacer(1, 0.15*inch))
        
        # Process each block
        for block in data['blocks']:
            # Helper to strip any HTML tags from user-supplied values
            def _strip_tags(s):
                if s is None:
                    return ''
                return re.sub(r'<[^>]+>', '', str(s))

            # Block header info
            header_data = [
                [Paragraph(_strip_tags(f"Center Code: {ExportManager.CENTER_CODE}"), normal_style),
                 Paragraph(_strip_tags(f"Date: {date_str}"), normal_style),
                 Paragraph(_strip_tags(f"Block Name: {block.get('blockName', '')}"), normal_style)],
            ]
            header_table = Table(header_data, colWidths=[2.2*inch, 2.2*inch, 2.2*inch])
            header_table.setStyle(TableStyle([
                ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 12),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ]))
            elements.append(header_table)
            elements.append(Spacer(1, 0.05*inch))
            
            # Block info
            # Use full PRN numbers and sanitize any user data
            info_data = [
                [Paragraph(_strip_tags(f"Total Students in Block: {block.get('totalStudents', '')}"), normal_style), 
                 Paragraph(_strip_tags(f"PRN No.: From {block.get('prnFrom', '')} to {block.get('prnTo', '')}"), normal_style)],
            ]
            info_table = Table(info_data, colWidths=[3.3*inch, 3.3*inch])
            info_table.setStyle(TableStyle([
                ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 0), (-1, -1), 12),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ]))
            elements.append(info_table)
            elements.append(Spacer(1, 0.05*inch))
            
            # Class and Branch info
            class_data = [
                [Paragraph(_strip_tags(f"Class: {block.get('year', '')} Year"), normal_style), 
                 Paragraph(_strip_tags(f"Branch: {block.get('branch', '')}"), normal_style)],
            ]
            class_table = Table(class_data, colWidths=[3.3*inch, 3.3*inch])
            class_table.setStyle(TableStyle([
                ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 0), (-1, -1), 12),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ]))
            elements.append(class_table)
            elements.append(Spacer(1, 0.1*inch))
            
            # Create seating table (vertical single-column layout for Desk No and PRN pairs)
            desk_data = [['Desk No.', 'PRN No.', 'Desk No.', 'PRN No.', 'Desk No.', 'PRN No.']]
            
            # Fill in student data in 3-column layout (2 pairs per row, 3 pairs across)
            students = block['students']
            for i in range(0, len(students), 3):
                row = []
                for j in range(3):
                    if i + j < len(students):
                        student = students[i + j]
                        row.append(str(student['deskNo']))
                        row.append(_strip_tags(student.get('prn', '')))
                    else:
                        row.append('')
                        row.append('')
                desk_data.append(row)
            
            desk_table = Table(desk_data, colWidths=[0.6*inch, 1.9*inch, 0.6*inch, 1.9*inch, 0.6*inch, 1.9*inch])
            desk_table.setStyle(TableStyle([
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 11),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
                ('TOPPADDING', (0, 0), (-1, -1), 6),
            ]))
            elements.append(desk_table)
            elements.append(Spacer(1, 0.3*inch))
            
            # Add page break between blocks if there are more
            if block != data['blocks'][-1]:
                elements.append(PageBreak())
        
        # Build PDF
        doc.build(elements)
        buffer.seek(0)
        return buffer
    
    @staticmethod
    def generate_excel(data):
        """Generate Excel with seating arrangements"""
        wb = Workbook()
        # Remove default sheet created by openpyxl
        if wb.active is not None:
            wb.remove(wb.active)

        # Font size to use in Excel
        excel_header_size = 12
        excel_body_size = 10

        for block_idx, block in enumerate(data.get('blocks', [])):
            ws = wb.create_sheet(title=f"Block-{block_idx + 1}")

            # Set column widths (wider to accommodate larger font)
            ws.column_dimensions['A'].width = 20
            ws.column_dimensions['B'].width = 40
            ws.column_dimensions['C'].width = 20
            ws.column_dimensions['D'].width = 40
            ws.column_dimensions['E'].width = 20
            ws.column_dimensions['F'].width = 40

            row_num = 1

            # Institute title
            ws.merge_cells(f'A{row_num}:F{row_num}')
            title_cell = ws[f'A{row_num}']
            title_cell.value = ExportManager.INSTITUTE_NAME
            # Institute title uses a slightly smaller size (20) per user request
            title_cell.font = Font(bold=True, size=20)
            title_cell.alignment = Alignment(horizontal='center', vertical='center')
            row_num += 2

            # Exam title and date
            now = datetime.now()
            month = now.month
            year = now.year
            # Use same mapping as PDF
            if month in [11, 12, 1]:
                exam_label = "Winter"
            elif month in [4, 5, 6, 7]:
                exam_label = "Summer"
            else:
                exam_label = ""
            exam_full = f"{exam_label + ' ' if exam_label else ''}Examination {year}"
            date_str = now.strftime('%A, %d %B %Y')

            ws.merge_cells(f'A{row_num}:F{row_num}')
            exam_cell = ws[f'A{row_num}']
            exam_cell.value = exam_full
            exam_cell.font = Font(bold=True, size=16)
            exam_cell.alignment = Alignment(horizontal='center', vertical='center')
            row_num += 2

            # Header info
            ws.merge_cells(f'A{row_num}:B{row_num}')
            ws[f'A{row_num}'].value = f"Center Code: {ExportManager.CENTER_CODE}"
            ws[f'A{row_num}'].font = Font(bold=True, size=excel_header_size)

            ws.merge_cells(f'C{row_num}:D{row_num}')
            ws[f'C{row_num}'].value = f"Date: {date_str}"
            ws[f'C{row_num}'].font = Font(bold=True, size=excel_header_size)

            ws.merge_cells(f'E{row_num}:F{row_num}')
            ws[f'E{row_num}'].value = f"Block Name: {block.get('blockName', '')}"
            ws[f'E{row_num}'].font = Font(bold=True, size=excel_header_size)
            row_num += 1

            # Student count and PRN range
            ws.merge_cells(f'A{row_num}:B{row_num}')
            ws[f'A{row_num}'].value = f"Total Students in Block: {block.get('totalStudents', '')}"
            ws[f'A{row_num}'].font = Font(bold=True, size=excel_body_size)

            ws.merge_cells(f'C{row_num}:F{row_num}')
            ws[f'C{row_num}'].value = f"PRN No.: From {block.get('prnFrom', '')} to {block.get('prnTo', '')}"
            ws[f'C{row_num}'].font = Font(bold=True, size=excel_body_size)
            row_num += 1

            # Class and Branch
            ws.merge_cells(f'A{row_num}:B{row_num}')
            ws[f'A{row_num}'].value = f"Class: {block.get('year', '')} Year"
            ws[f'A{row_num}'].font = Font(bold=True, size=excel_body_size)

            ws.merge_cells(f'C{row_num}:D{row_num}')
            ws[f'C{row_num}'].value = f"Branch: {block.get('branch', '')}"
            ws[f'C{row_num}'].font = Font(bold=True, size=excel_body_size)
            row_num += 2

            # Seating table header (vertical single-column layout)
            headers = ['Desk No.', 'PRN No.', 'Desk No.', 'PRN No.', 'Desk No.', 'PRN No.']
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=row_num, column=col_idx)
                cell.value = header
                cell.font = Font(bold=True, color="FFFFFF", size=excel_header_size)
                cell.fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")
                cell.alignment = Alignment(horizontal='center', vertical='center')
            row_num += 1
            
            # Add student data (3-column layout: pairs of Desk/PRN)
            border = Border(left=Side(style='thin'), right=Side(style='thin'),
                           top=Side(style='thin'), bottom=Side(style='thin'))
            
            students = block.get('students', [])
            for i in range(0, len(students), 3):
                col = 1
                for j in range(3):
                    if i + j < len(students):
                        student = students[i + j]
                        # Desk No
                        desk_cell = ws.cell(row=row_num, column=col)
                        desk_cell.value = student.get('deskNo', '')
                        desk_cell.font = Font(size=excel_body_size)
                        desk_cell.alignment = Alignment(horizontal='center', vertical='center')
                        desk_cell.border = border
                        
                        # PRN No
                        prn_cell = ws.cell(row=row_num, column=col + 1)
                        prn_cell.value = str(student.get('prn', ''))
                        prn_cell.font = Font(size=excel_body_size)
                        prn_cell.alignment = Alignment(horizontal='center', vertical='center')
                        prn_cell.border = border
                        
                        col += 2
                row_num += 1

            # Add student data (vertical single-column layout)
            border = Border(left=Side(style='thin'), right=Side(style='thin'),
                          top=Side(style='thin'), bottom=Side(style='thin'))

            students = block.get('students', [])
            for student in students:
                # Desk No (Column A)
                desk_cell = ws.cell(row=row_num, column=1)
                desk_cell.value = student.get('deskNo', '')
                desk_cell.font = Font(size=excel_body_size)
                desk_cell.alignment = Alignment(horizontal='center', vertical='center')
                desk_cell.border = border

                # PRN No (Column B)
                prn_cell = ws.cell(row=row_num, column=2)
                prn_cell.value = str(student.get('prn', ''))
                prn_cell.font = Font(size=excel_body_size)
                prn_cell.alignment = Alignment(horizontal='center', vertical='center')
                prn_cell.border = border
                
                row_num += 1

        # Save to buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer
